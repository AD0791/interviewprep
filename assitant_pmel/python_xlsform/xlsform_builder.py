#!/usr/bin/env python3
"""
Construire un XLSForm par programme, sans ouvrir Excel.
Usage : python3 xlsform_builder.py            -> genere exemple_genere.xlsx
Dependances : openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date

class XLSForm:
    """Petit constructeur : on empile des questions, on ecrit le classeur."""

    def __init__(self, titre, form_id, langue="francais (fr)"):
        self.titre, self.form_id, self.langue = titre, form_id, langue
        self.survey, self.choices = [], []
        self.version = date.today().strftime("%Y%m%d") + "01"

    # -- questions -------------------------------------------------------
    def q(self, type_, name, label="", **options):
        self.survey.append(dict(type=type_, name=name, label=label, **options))
        return self

    def metadonnees(self):
        for t in ("start", "end", "today", "deviceid"):
            self.q(t, t)
        return self

    def groupe(self, name, label, apparence="field-list"):
        self.q("begin_group", name, label, appearance=apparence)
        return self

    def fin_groupe(self, name):
        self.q("end_group", f"{name}_fin")
        return self

    def repetition(self, name, label, compte=None):
        opts = {"repeat_count": compte} if compte else {}
        self.q("begin_repeat", name, label, **opts)
        return self

    def fin_repetition(self, name):
        self.q("end_repeat", f"{name}_fin")
        return self

    def calcul(self, name, expression):
        self.q("calculate", name, calculation=expression)
        return self

    # -- listes de choix -------------------------------------------------
    def liste(self, list_name, items, colonne_filtre=None):
        """items : [(name, label)] ou [(name, label, valeur_du_filtre)]"""
        for item in items:
            ligne = {"list_name": list_name, "name": item[0], "label": item[1]}
            if colonne_filtre and len(item) > 2:
                ligne[colonne_filtre] = item[2]
            self.choices.append(ligne)
        return self

    # -- ecriture --------------------------------------------------------
    def ecrire(self, chemin):
        wb = Workbook()
        entete = Font(bold=True, color="FFFFFF")
        fond = PatternFill("solid", fgColor="4F6228")

        def feuille(ws, lignes, ordre_prioritaire=()):
            colonnes = list(ordre_prioritaire)
            for l in lignes:
                for k in l:
                    if k not in colonnes:
                        colonnes.append(k)
            ws.append(colonnes)
            for c in ws[1]:
                c.font, c.fill = entete, fond
            for l in lignes:
                ws.append([l.get(c, "") for c in colonnes])
            for i, c in enumerate(colonnes, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(12, min(38, len(c) + 6))
            ws.freeze_panes = "A2"

        ws = wb.active; ws.title = "survey"
        feuille(ws, self.survey, ("type", "name", "label"))
        feuille(wb.create_sheet("choices"), self.choices, ("list_name", "name", "label"))
        feuille(wb.create_sheet("settings"),
                [{"form_title": self.titre, "form_id": self.form_id,
                  "version": self.version, "default_language": self.langue}],
                ("form_title", "form_id", "version", "default_language"))
        wb.save(chemin)
        return chemin


# ---------------------------------------------------------------- exemple
def exemple_suivi_cantine():
    f = XLSForm("Suivi quotidien de la cantine", "pa_cantine_jour")
    f.metadonnees()
    f.liste("regions", [("ouest", "Ouest"), ("sud", "Sud"),
                        ("centre", "Centre"), ("artibonite", "Artibonite")])
    f.liste("ecoles", [("ec001", "Ecole Ganthier I", "ouest"),
                       ("ec004", "Ecole Puits-Sale", "sud"),
                       ("ec007", "Ecole Mirebalais", "centre"),
                       ("ec009", "Ecole Gonaives", "artibonite")], colonne_filtre="region")
    f.liste("oui_non", [("oui", "Oui"), ("non", "Non")])
    f.liste("motifs", [("rupture", "Rupture de stock"), ("transport", "Probleme de transport"),
                       ("securite", "Insecurite"), ("autre", "Autre")])

    f.groupe("entete", "Identification")
    f.q("date", "date_jour", "Date du service", required="yes",
        constraint=".<=today()", constraint_message="Date future impossible")
    f.q("select_one regions", "region", "Departement", required="yes")
    f.q("select_one ecoles", "ecole", "Ecole", required="yes", choice_filter="region=${region}")
    f.fin_groupe("entete")

    f.q("integer", "presents", "Eleves presents aujourd'hui", required="yes",
        constraint=".>=0 and .<=500", constraint_message="Valeur attendue entre 0 et 500")
    f.q("select_one oui_non", "service_assure", "Le repas a-t-il ete servi ?", required="yes")
    f.q("integer", "repas", "Nombre de repas servis", required="yes",
        relevant="${service_assure}='oui'",
        constraint=".<=${presents}",
        constraint_message="Impossible : plus de repas que d'eleves presents")
    f.q("select_one motifs", "motif", "Pourquoi le repas n'a-t-il pas ete servi ?",
        required="yes", relevant="${service_assure}='non'")
    f.calcul("couverture", "if(${presents}>0, round(${repas} div ${presents}*100,1), 0)")
    f.q("note", "recap", "Couverture du jour : ${couverture} % des eleves presents")
    f.q("image", "photo", "Photo du registre de cantine")
    return f


if __name__ == "__main__":
    chemin = exemple_suivi_cantine().ecrire("exemple_genere.xlsx")
    print("Genere :", chemin)

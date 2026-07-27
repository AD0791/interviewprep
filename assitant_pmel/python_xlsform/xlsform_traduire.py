#!/usr/bin/env python3
"""
Ajouter ou completer une langue dans un XLSForm existant.
Usage : python3 xlsform_traduire.py source.xlsx sortie.xlsx "kreyol (ht)"
        (si le formulaire est monolingue, les colonnes label/hint/constraint_message
         sont converties en label::langue_par_defaut, etc.)
Dependances : openpyxl
"""
import sys
from openpyxl import load_workbook

TRADUISIBLES = ("label", "hint", "constraint_message", "required_message", "image", "audio")

def colonnes(ws):
    return [str(c.value).strip() if c.value else "" for c in ws[1]]

def traduire(source, sortie, nouvelle_langue, langue_source=None):
    wb = load_workbook(source)
    settings = wb["settings"]
    cols_set = colonnes(settings)
    defaut = None
    if "default_language" in cols_set:
        defaut = settings.cell(row=2, column=cols_set.index("default_language") + 1).value
    langue_source = langue_source or defaut or "francais (fr)"

    for nom_feuille in ("survey", "choices"):
        if nom_feuille not in wb.sheetnames:
            continue
        ws = wb[nom_feuille]
        cols = colonnes(ws)

        # 1. monolingue -> multilingue : on renomme label en label::<langue_source>
        for i, c in enumerate(cols):
            if c in TRADUISIBLES:
                ws.cell(row=1, column=i + 1).value = f"{c}::{langue_source}"
        cols = colonnes(ws)

        # 2. pour chaque colonne existante dans la langue source, creer l'equivalent
        a_creer = []
        for c in cols:
            if "::" in c:
                base, lang = c.split("::", 1)
                if lang == langue_source and f"{base}::{nouvelle_langue}" not in cols:
                    a_creer.append((base, cols.index(c) + 1))

        for base, col_src in a_creer:
            nouvelle_col = ws.max_column + 1
            ws.cell(row=1, column=nouvelle_col).value = f"{base}::{nouvelle_langue}"
            for r in range(2, ws.max_row + 1):
                valeur = ws.cell(row=r, column=col_src).value
                # pre-remplissage : on recopie le texte source prefixe, a traduire ensuite
                ws.cell(row=r, column=nouvelle_col).value = (
                    f"[A TRADUIRE] {valeur}" if valeur else None)

    wb.save(sortie)
    return sortie

def rapport_traduction(chemin):
    """Liste ce qui reste a traduire."""
    wb = load_workbook(chemin, data_only=True)
    restant = []
    for nom in ("survey", "choices"):
        if nom not in wb.sheetnames:
            continue
        ws = wb[nom]
        cols = colonnes(ws)
        for i, c in enumerate(cols):
            if "::" not in c:
                continue
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=i + 1).value
                if v and str(v).startswith("[A TRADUIRE]"):
                    restant.append((nom, r, c, str(v)[:45]))
    return restant

if __name__ == "__main__":
    src, out = sys.argv[1], sys.argv[2]
    langue = sys.argv[3] if len(sys.argv) > 3 else "kreyol (ht)"
    traduire(src, out, langue)
    reste = rapport_traduction(out)
    print(f"Ecrit : {out}")
    print(f"Chaines a traduire : {len(reste)}")
    for f, r, c, v in reste[:8]:
        print(f"  {f} ligne {r} [{c}] {v}")

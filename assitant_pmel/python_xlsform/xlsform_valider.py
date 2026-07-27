#!/usr/bin/env python3
"""
Valider en lot des XLSForm avec pyxform (conversion en XForm).
Usage : python3 xlsform_valider.py dossier_ou_fichiers...
Dependances : pyxform
"""
import glob, os, sys, tempfile
from pyxform.xls2xform import xls2xform_convert

def est_xlsform(chemin):
    """Un XLSForm doit avoir une feuille 'survey'. Sinon c'est un classeur de donnees."""
    from openpyxl import load_workbook
    try:
        return "survey" in load_workbook(chemin, read_only=True).sheetnames
    except Exception:
        return False

def valider(chemin):
    sortie = os.path.join(tempfile.gettempdir(), os.path.basename(chemin) + ".xml")
    try:
        avertissements = xls2xform_convert(chemin, sortie, validate=False)
        return True, avertissements or [], sortie
    except Exception as e:
        return False, [str(e)], None

if __name__ == "__main__":
    cibles = []
    for a in sys.argv[1:]:
        cibles.extend(glob.glob(os.path.join(a, "*.xlsx")) if os.path.isdir(a) else [a])
    cibles = [c for c in cibles if est_xlsform(c)] or cibles
    ok = 0
    for c in sorted(cibles):
        succes, messages, xml = valider(c)
        print(("[OK]   " if succes else "[ECHEC]") + " " + os.path.basename(c))
        for m in messages[:3]:
            print("        " + str(m)[:110])
        if succes:
            ok += 1
    print(f"\n{ok}/{len(cibles)} formulaire(s) valide(s).")
    sys.exit(0 if ok == len(cibles) else 1)

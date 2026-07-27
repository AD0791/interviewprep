#!/usr/bin/env python3
"""
Audit d'un fichier XLSForm : structure, references, contraintes, traductions.
Usage : python3 xlsform_audit.py mon_formulaire.xlsx
Dependances : openpyxl
"""
import re, sys
from collections import Counter, defaultdict
from openpyxl import load_workbook

NAME_RE = re.compile(r"^[a-z][a-z0-9_]*$")
REF_RE  = re.compile(r"\$\{([^}]+)\}")
META = {"start","end","today","deviceid","username","phonenumber","subscriberid",
        "simserial","audit","begin_group","end_group","begin_repeat","end_repeat"}

def lire(chemin):
    wb = load_workbook(chemin, data_only=True)
    feuilles = {}
    for nom in ("survey", "choices", "settings"):
        if nom not in wb.sheetnames:
            feuilles[nom] = []
            continue
        ws = wb[nom]
        lignes = list(ws.values)
        if not lignes:
            feuilles[nom] = []
            continue
        entetes = [str(c).strip() if c is not None else "" for c in lignes[0]]
        feuilles[nom] = [
            {h: ("" if v is None else str(v).strip()) for h, v in zip(entetes, ligne)}
            for ligne in lignes[1:] if any(v is not None for v in ligne)
        ]
    return feuilles

def auditer(chemin):
    f = lire(chemin)
    survey, choices, settings = f["survey"], f["choices"], f["settings"]
    problemes, infos = [], []

    # --- structure ---
    types = Counter(r.get("type","").split()[0] for r in survey if r.get("type"))
    noms  = [r.get("name","") for r in survey if r.get("name")]
    listes_definies = {c.get("list_name","") for c in choices if c.get("list_name")}
    listes_utilisees = {r["type"].split()[1] for r in survey
                        if r.get("type","").startswith(("select_one ","select_multiple "))}

    infos.append(f"Questions : {len(survey)}")
    infos.append(f"Listes de choix definies : {len(listes_definies)} | utilisees : {len(listes_utilisees)}")
    infos.append("Types : " + ", ".join(f"{t}={n}" for t, n in types.most_common()))
    infos.append(f"Obligatoires : {sum(1 for r in survey if r.get('required','').lower() in ('yes','true','oui'))}")
    infos.append(f"Contraintes : {sum(1 for r in survey if r.get('constraint'))}")
    infos.append(f"Conditions (relevant) : {sum(1 for r in survey if r.get('relevant'))}")
    infos.append(f"Calculs : {sum(1 for r in survey if r.get('calculation'))}")

    # --- nommage ---
    for n in noms:
        if not NAME_RE.match(n):
            problemes.append(f"NOM INVALIDE : '{n}' (attendu : minuscules, chiffres, underscore, debut par une lettre)")
    for n, c in Counter(noms).items():
        if c > 1:
            problemes.append(f"NOM DUPLIQUE : '{n}' apparait {c} fois")

    # --- listes ---
    for l in sorted(listes_utilisees - listes_definies):
        problemes.append(f"LISTE MANQUANTE : '{l}' est utilisee dans survey mais absente de choices")
    for l in sorted(listes_definies - listes_utilisees):
        problemes.append(f"LISTE ORPHELINE : '{l}' est definie mais jamais utilisee")

    # --- references ${...} ---
    connus = set(noms) | META
    for i, r in enumerate(survey, start=2):
        for col in ("relevant","constraint","calculation","choice_filter","repeat_count","default"):
            for ref in REF_RE.findall(r.get(col,"")):
                if ref not in connus:
                    problemes.append(f"REFERENCE INCONNUE ligne {i} ({col}) : ${{{ref}}}")
        for col in [c for c in r if c.startswith("label")]:
            for ref in REF_RE.findall(r.get(col,"")):
                if ref not in connus:
                    problemes.append(f"REFERENCE INCONNUE ligne {i} (label) : ${{{ref}}}")

    # --- contraintes sans message ---
    for i, r in enumerate(survey, start=2):
        if r.get("constraint") and not any(v for k, v in r.items() if k.startswith("constraint_message")):
            problemes.append(f"CONTRAINTE SANS MESSAGE ligne {i} : {r.get('name')}")

    # --- division ---
    for i, r in enumerate(survey, start=2):
        calc = r.get("calculation","")
        if re.search(r"\}\s*/\s*\$", calc):
            problemes.append(f"DIVISION AVEC '/' ligne {i} : utiliser 'div' en XPath -> {r.get('name')}")
        if calc.count("div") and "if(" not in calc and "NULLIF" not in calc:
            problemes.append(f"DIVISION NON PROTEGEE ligne {i} : risque de division par zero -> {r.get('name')}")

    # --- traductions ---
    langues = defaultdict(set)
    for r in survey:
        for k in r:
            if "::" in k:
                col, lang = k.split("::", 1)
                langues[lang].add(col)
    if langues:
        toutes = set().union(*langues.values())
        infos.append("Langues : " + ", ".join(sorted(langues)))
        for lang, cols in langues.items():
            for manquante in sorted(toutes - cols):
                problemes.append(f"TRADUCTION INCOMPLETE : la langue '{lang}' n'a pas de colonne '{manquante}'")
        for lang in langues:
            vides = [r.get("name") for r in survey
                     if r.get(f"label::{lang}", None) == "" and r.get("type","") not in META
                     and any(r.get(f"label::{l}") for l in langues if l != lang)]
            if vides:
                problemes.append(f"LIBELLES VIDES en '{lang}' : {', '.join(v for v in vides[:5] if v)}")

    # --- settings ---
    if not settings:
        problemes.append("SETTINGS ABSENT : ajouter form_title, form_id et version")
    else:
        s = settings[0]
        for champ in ("form_title","form_id","version"):
            if not s.get(champ):
                problemes.append(f"SETTINGS INCOMPLET : '{champ}' manquant")

    return infos, problemes

if __name__ == "__main__":
    for chemin in sys.argv[1:]:
        print("=" * 70)
        print(chemin)
        print("=" * 70)
        infos, problemes = auditer(chemin)
        for i in infos:
            print("  " + i)
        print()
        if problemes:
            print(f"  {len(problemes)} point(s) d'attention :")
            for p in problemes:
                print("   - " + p)
        else:
            print("  Aucun probleme detecte.")
        print()
